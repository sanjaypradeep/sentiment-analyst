import pandas as pd

from . import preprocess
# from scripts.preprocess import PreProcess


from sklearn.model_selection import train_test_split
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.externals import joblib
from sklearn import svm
from sklearn.naive_bayes import MultinomialNB

from . import constants
from copy import copy,deepcopy

def pre_process_data(data):

	#the pre processing part
	column_name = data.columns[0]
	# data=df
	pre_processor = preprocess.PreProcess(data, column_name)	

	data = pre_processor.clean_html()
	data = pre_processor.remove_non_ascii()
	data = pre_processor.remove_spaces()
	data = pre_processor.remove_punctuation()
	data = pre_processor.stemming()
	data = pre_processor.lemmatization()
	data = pre_processor.stop_words()

	return data



def test_model(test_text,test_file_name,test_reference_file,models_list):
	#returns dataframe with label, status as True or False, 
	
	print("text is ",test_text," test file is ",test_file_name," reference file is ",test_reference_file," list of models ",models_list)
	print("constant = ",constants.vectorlibs_location,constants.trained_models_location)
	identifier=test_reference_file

	test_is_a_file=False
	#this flag becomes true if test data is a csv file
	#also enabbles the result dataframe to be written to 
	#another csv file
	#you give csv, you get csv

	if test_file_name is None and test_text!="":
		#deal with the string in text box
		print("Getting text",len(test_text))
		if len(test_text) < 20:
			print("Please provide input large enough, Classifier can understand :)")
			return None,None,False
		else:
			print("Generating the dataframe from text")
			d = {'Text': [test_text]}
			df = pd.DataFrame(data=d)
			print("Done")
			print(df.head())

	#need to consider file
	elif test_file_name != None and  test_text=="":
		print("test file name is ",test_file_name)
		test_is_a_file=True
		df=pd.read_csv(test_file_name)
		print(df.head(5))


	else:
		print("What am i doing here")
		return None,None,False


	#keep a backup of df

	df_bkp=deepcopy(df)



	#now work on the dataframe df
	vectorizer=constants.vectorlibs_location+identifier+'_vectorizer.pkl'
	tfidf_transformer = joblib.load(vectorizer)
	print("Loaded vectorizer ",vectorizer)
	print(vectorizer)

	#pre process data 
	data=pre_process_data(df)
	print("After pre processing")
	print(data.head(5))
	
	col1=data.columns[0]
	data_check = tfidf_transformer.transform(data[col1])
	print("The data after pre process and transform is")
	print(data.head(5))
	
	print("data check type is :",type(data_check))
	print("Backup looks like this")
	print(df_bkp.head(5))
	
	


	for model_name in models_list:
		model_file= str(constants.trained_models_location)+str(identifier)+"_"+str(model_name)+'.pkl'
		model=joblib.load(model_file)
		print("Loaded ",model_file)
		output=model.predict(data_check)
		print("After running model")
		print(output)
		df_bkp[model_name]=output
		# for i in range(0,len(output)):
		# 	result[i][model_name]=output[i]


	print("After testing result is")
	print(df_bkp.head())
	if test_is_a_file:
		#write to outputs/results
		print("will write output df to the following location",constants.output_results_location)
		df_bkp.to_csv(constants.output_results_location+str(test_file_name)+"_results.csv")

		#hold on, we will also write it to an excel

		#add the sheet to a list of sheets
		result_df_list=[]
		result_df_list.append(df_bkp)

		xls_name=constants.output_results_location+ "result.xlsx"
		
		from openpyxl import Workbook
		from openpyxl.utils.dataframe import dataframe_to_rows
		
		# # create Workbook object
		wb=Workbook()
		print("loaded some workbook")
		sheetnames=["result"]

		for n,dataf in enumerate(result_df_list):
			print("n is ",n)
			print(dataf.head())
			ws=wb.create_sheet(sheetnames[n])
			for r in dataframe_to_rows(dataf, index=True, header=True):
				ws.append(r)

		print("going to save at",xls_name)

		#remove the default sheet called "sheet"
		wb.remove(wb.get_sheet_by_name('Sheet'))
		wb.save(xls_name)

		print("written to xls")
		return df_bkp,xls_name,True



	
	return df_bkp,None,True
